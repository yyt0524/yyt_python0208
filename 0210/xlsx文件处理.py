# -*- coding: UTF-8 -*-
import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 'yyt'
ws.append([1,2,3])
ws['A2'] = datetime.datetime.now()
wb.save('yyt_excel.xlsx')